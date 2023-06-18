from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os

folder_name = "GENERATEDIDs"
if not os.path.exists(folder_name):
    os.makedirs(folder_name)

workbook = load_workbook('student_information.xlsx')
sheet = workbook.active  

for row in sheet.iter_rows(min_row=2): 

    id_card = Image.new('RGB', (800, 400), (255, 255, 255))
    draw = ImageDraw.Draw(id_card)

    logo = Image.open("logo.webp")  
    logo = logo.resize((115,112))
    id_card.paste(logo, (25, 0))

    institute_name = "Maharaja Surajmal Institute"
    about = "(Affiliated to GGSIP University,Dwarka,Delhi)"
    address = "C-4, Janakpuri, New Delhi-110058"
    phone_number = "Tel. : 011-45656183, 011-45037193"
    font_name = ImageFont.truetype('DejaVuSans', 35) 
    font_info = ImageFont.truetype('DejaVuSans', 12)
    draw.text((160, 10), institute_name, font=font_name, fill=(0,0,255))
    draw.text((266, 50), about, font=font_info, fill=(0, 255, 0	))
    draw.text((300, 65), address, font=font_info, fill=(255, 0, 0	))
    draw.text((294, 80), phone_number, font=font_info, fill=(255, 0, 0	))

    draw.line([(0,105), (800,105)], fill=(0, 255, 0	), width=2)

    name = row[0].value
    fathers_name = row[1].value
    date_of_birth = row[2].value
    cet_rank = row[3].value
    phone_number = row[4].value
    course_batch = row[5].value
    photograph_path = './photos/'+row[6].value
    digital_sign_path ='./sign/'+row[7].value
    batch = row[9].value
    address = row[8].value

    photo = Image.open(photograph_path)
    signature = Image.open(digital_sign_path)

    font = ImageFont.truetype('DejaVuSans', 20)  
    draw.text((175, 125), f'Name: {name}', fill='black', font=font)
    draw.text((175, 150), f'Father\'s Name: {fathers_name}', fill='black', font=font)
    draw.text((175, 175), f'Date of Birth: {date_of_birth}', fill='black', font=font)
    draw.text((550, 175), f'CET Rank: {cet_rank}', fill='black', font=font)
    draw.text((175, 225), f'Phone Number: {phone_number}', fill='black', font=font)
    draw.text((175, 200), f'Course: {course_batch}', fill='black', font=font)
    draw.text((550, 200), f'Batch: {batch}', fill='black', font=font)
    draw.text((175, 250), f'Res Address: {address}', fill='black', font=font)

    photo = photo.resize((125, 120))
    id_card.paste(photo, (25, 120))

    signature = signature.resize((125, 50))
    id_card.paste(signature,(25,250) )

    id_card.save(os.path.join(folder_name, f'{cet_rank}_id_card.jpg'))
print('ID card generation complete!')
